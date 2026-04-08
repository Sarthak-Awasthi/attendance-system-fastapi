from __future__ import annotations

import asyncio
from datetime import datetime

from openpyxl import load_workbook

from core.config import settings
from services.excel_service import append_attendance_row, initialize_worksheet


async def main() -> None:
    course = "CS301"
    sheet = f"phase8_{datetime.now().strftime('%H%M%S')}"
    await initialize_worksheet(course, sheet)

    async def write_one(i: int) -> None:
        await append_attendance_row(
            course,
            sheet,
            [f"C{i:04d}", datetime.now().isoformat(), "phase8-session", f"10.0.0.{i}", 1, "CONCUR"],
        )

    await asyncio.gather(*(write_one(i) for i in range(1, 41)))

    workbook = load_workbook(settings.excel_data_dir / f"{course}.xlsx")
    rows = workbook[sheet].max_row
    assert rows == 41, f"Expected 41 rows including header, got {rows}"
    print("phase8_concurrency_test.py passed")


if __name__ == "__main__":
    asyncio.run(main())

