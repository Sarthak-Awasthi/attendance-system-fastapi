from __future__ import annotations

import asyncio
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from core.config import settings

ROLL_HEADER = "RollNo."
course_locks: defaultdict[str, asyncio.Lock] = defaultdict(asyncio.Lock)


async def get_next_worksheet_name(course_code: str) -> str:
    lock = course_locks[course_code]
    async with lock:
        return await asyncio.to_thread(_get_next_worksheet_name_sync, course_code)


def _get_next_worksheet_name_sync(course_code: str) -> str:
    file_path = _course_file_path(course_code)
    date_prefix = datetime.now().strftime("%Y-%m-%d")
    pattern = re.compile(rf"^{re.escape(date_prefix)}_S(\d+)$")

    highest = 0
    if file_path.exists():
        workbook = load_workbook(file_path)
        for name in workbook.sheetnames:
            match = pattern.match(name)
            if match:
                highest = max(highest, int(match.group(1)))

    return f"{date_prefix}_S{highest + 1}"


def _course_file_path(course_code: str) -> Path:
    settings.excel_data_dir.mkdir(parents=True, exist_ok=True)
    return settings.excel_data_dir / f"{course_code}.xlsx"


def _setup_sheet(sheet) -> None:
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        sheet.delete_rows(1)
    if sheet.max_row == 0:
        sheet.append([ROLL_HEADER])
    if sheet["A1"].value != ROLL_HEADER:
        sheet["A1"] = ROLL_HEADER
    sheet["A1"].font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 18


def _ensure_date_column(sheet, attendance_date: str) -> int:
    for col in range(2, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col).value or "") == attendance_date:
            return col

    new_col = max(2, sheet.max_column + 1)
    date_cell = sheet.cell(row=1, column=new_col, value=attendance_date)
    date_cell.font = Font(bold=True)
    sheet.column_dimensions[date_cell.column_letter].width = 14

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=new_col).value is None:
            sheet.cell(row=row, column=new_col, value=0)

    return new_col


def _ensure_roll_row(sheet, roll_number: str) -> int:
    normalized_roll = roll_number.upper().strip()
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value or "").upper().strip() == normalized_roll:
            return row

    new_row = max(2, sheet.max_row + 1)
    sheet.cell(row=new_row, column=1, value=normalized_roll)
    for col in range(2, sheet.max_column + 1):
        sheet.cell(row=new_row, column=col, value=0)
    return new_row


async def initialize_worksheet(course_code: str, worksheet_name: str) -> None:
    lock = course_locks[course_code]
    async with lock:
        await asyncio.to_thread(_initialize_worksheet_sync, course_code, worksheet_name)


def _initialize_worksheet_sync(course_code: str, worksheet_name: str) -> None:
    file_path = _course_file_path(course_code)
    if file_path.exists():
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

    if worksheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=worksheet_name)
        _setup_sheet(sheet)
    else:
        _setup_sheet(workbook[worksheet_name])

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        default_sheet = workbook["Sheet"]
        if default_sheet.max_row == 1 and default_sheet["A1"].value is None:
            workbook.remove(default_sheet)

    workbook.save(file_path)


async def append_attendance_row(course_code: str, worksheet_name: str, row_data: Iterable) -> None:
    row = list(row_data)
    if len(row) < 2:
        raise ValueError("Attendance row must include at least roll number and date")

    roll_number = str(row[0])
    attendance_date = str(row[1])
    present = 1
    if len(row) > 5:
        present = 1 if int(row[5]) else 0
    elif len(row) > 2:
        present = 1 if int(row[2]) else 0

    await mark_attendance(course_code, worksheet_name, roll_number, attendance_date=attendance_date, present=present)


async def mark_attendance(
    course_code: str,
    worksheet_name: str,
    roll_number: str,
    attendance_date: str | None = None,
    present: int = 1,
) -> None:
    """Mark attendance for a student in Excel workbook.

    This function is thread-safe using per-course locks. It creates or updates
    the attendance record in the Excel file with a 1 (present) or 0 (absent).

    Args:
        course_code: Course code (used to find workbook)
        worksheet_name: Sheet name within workbook
        roll_number: Student's roll number
        attendance_date: Date in YYYY-MM-DD format; defaults to today
        present: 1 for present, 0 for absent (default: 1)

    Raises:
        FileNotFoundError: If course workbook doesn't exist
        ValueError: If worksheet doesn't exist
        IOError: If file I/O fails
    """
    lock = course_locks[course_code]
    async with lock:
        await asyncio.to_thread(
            _mark_attendance_sync,
            course_code,
            worksheet_name,
            roll_number,
            attendance_date or date.today().isoformat(),
            1 if present else 0,
        )


def _mark_attendance_sync(
    course_code: str,
    worksheet_name: str,
    roll_number: str,
    attendance_date: str,
    present: int,
) -> None:
    file_path = _course_file_path(course_code)
    if not file_path.exists():
        raise FileNotFoundError(f"Course workbook does not exist: {file_path}")

    workbook = load_workbook(file_path)
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet does not exist: {worksheet_name}")

    sheet = workbook[worksheet_name]
    _setup_sheet(sheet)
    date_col = _ensure_date_column(sheet, attendance_date)
    row_num = _ensure_roll_row(sheet, roll_number)
    sheet.cell(row=row_num, column=date_col, value=1 if present else 0)
    workbook.save(file_path)

