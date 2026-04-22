"""Tests for the Excel service."""
from __future__ import annotations

import asyncio
import shutil
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.config import settings
from services.excel_service import (
    _course_file_path,
    initialize_worksheet,
    mark_attendance,
)


@pytest.fixture(autouse=True)
def _use_test_data_dir(tmp_path: Path):
    """Use a temporary directory for Excel data in every test."""
    original = settings.excel_data_dir
    settings.excel_data_dir = tmp_path / "test_excel"
    settings.excel_data_dir.mkdir(parents=True, exist_ok=True)
    yield
    settings.excel_data_dir = original


class TestInitializeWorksheet:
    async def test_creates_file_and_sheet(self):
        await initialize_worksheet("TEST001", "2024-01-01_S1")
        file_path = _course_file_path("TEST001")
        assert file_path.exists()
        wb = load_workbook(file_path)
        assert "2024-01-01_S1" in wb.sheetnames
        sheet = wb["2024-01-01_S1"]
        assert sheet["A1"].value == "RollNo."

    async def test_idempotent(self):
        await initialize_worksheet("TEST001", "sheet1")
        await initialize_worksheet("TEST001", "sheet1")
        wb = load_workbook(_course_file_path("TEST001"))
        assert wb.sheetnames.count("sheet1") == 1

    async def test_removes_default_sheet(self):
        await initialize_worksheet("TEST001", "my_sheet")
        wb = load_workbook(_course_file_path("TEST001"))
        assert "Sheet" not in wb.sheetnames

    async def test_multiple_worksheets(self):
        await initialize_worksheet("TEST001", "S1")
        await initialize_worksheet("TEST001", "S2")
        wb = load_workbook(_course_file_path("TEST001"))
        assert "S1" in wb.sheetnames
        assert "S2" in wb.sheetnames


class TestMarkAttendance:
    async def test_marks_present(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "21CS1001", "2024-01-15", present=1)
        wb = load_workbook(_course_file_path("TEST001"))
        ws = wb["test_ws"]
        # Row 2 should have the roll number, and column 2 should have the date
        assert ws.cell(row=2, column=1).value == "21CS1001"
        assert ws.cell(row=1, column=2).value == "2024-01-15"
        assert ws.cell(row=2, column=2).value == 1

    async def test_marks_absent(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "21CS1001", "2024-01-15", present=0)
        wb = load_workbook(_course_file_path("TEST001"))
        assert wb["test_ws"].cell(row=2, column=2).value == 0

    async def test_duplicate_roll_updates_not_duplicates(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "21CS1001", "2024-01-15", present=1)
        await mark_attendance("TEST001", "test_ws", "21CS1001", "2024-01-15", present=0)
        wb = load_workbook(_course_file_path("TEST001"))
        ws = wb["test_ws"]
        # Should still only have 2 rows (header + 1 student)
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == 0

    async def test_multiple_students(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "STU001", "2024-01-15", present=1)
        await mark_attendance("TEST001", "test_ws", "STU002", "2024-01-15", present=1)
        await mark_attendance("TEST001", "test_ws", "STU003", "2024-01-15", present=0)
        wb = load_workbook(_course_file_path("TEST001"))
        ws = wb["test_ws"]
        assert ws.max_row == 4  # header + 3 students

    async def test_multiple_dates(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "STU001", "2024-01-15", present=1)
        await mark_attendance("TEST001", "test_ws", "STU001", "2024-01-16", present=0)
        wb = load_workbook(_course_file_path("TEST001"))
        ws = wb["test_ws"]
        assert ws.cell(row=1, column=2).value == "2024-01-15"
        assert ws.cell(row=1, column=3).value == "2024-01-16"
        assert ws.cell(row=2, column=2).value == 1
        assert ws.cell(row=2, column=3).value == 0

    async def test_roll_number_case_insensitive(self):
        await initialize_worksheet("TEST001", "test_ws")
        await mark_attendance("TEST001", "test_ws", "abc123", "2024-01-15", present=1)
        await mark_attendance("TEST001", "test_ws", "ABC123", "2024-01-15", present=0)
        wb = load_workbook(_course_file_path("TEST001"))
        ws = wb["test_ws"]
        # Should have updated same row, not created a new one
        assert ws.max_row == 2

    async def test_nonexistent_workbook_raises(self):
        with pytest.raises(FileNotFoundError):
            await mark_attendance("NOSUCHCOURSE", "test_ws", "STU001", "2024-01-15")

    async def test_nonexistent_worksheet_raises(self):
        await initialize_worksheet("TEST001", "real_sheet")
        with pytest.raises(ValueError, match="Worksheet does not exist"):
            await mark_attendance("TEST001", "fake_sheet", "STU001", "2024-01-15")


class TestConcurrentWrites:
    async def test_concurrent_attendance_writes(self):
        course = "CONCURRENT"
        sheet = f"concurrent_{datetime.now().strftime('%H%M%S')}"
        await initialize_worksheet(course, sheet)

        async def write_one(i: int) -> None:
            await mark_attendance(
                course, sheet, f"S{i:04d}",
                attendance_date="2024-06-01", present=1,
            )

        await asyncio.gather(*(write_one(i) for i in range(1, 31)))
        wb = load_workbook(_course_file_path(course))
        rows = wb[sheet].max_row
        assert rows == 31, f"Expected 31 rows (header + 30 students), got {rows}"
